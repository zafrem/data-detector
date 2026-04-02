"""File storage adapter for scanning CSV, JSON, Parquet, and Excel files.

CSV and JSON support uses stdlib only. Excel requires openpyxl,
Parquet requires pyarrow (both optional).

Install: pip install data-detector[file-storage]
"""

import csv
import fnmatch
import json
import logging
from pathlib import Path
from typing import List, Optional

from datadetector.resource_adapter import ResourceAdapter
from datadetector.resource_models import (
    ContainerInfo,
    ContainerType,
    DataResource,
    FieldInfo,
)

logger = logging.getLogger(__name__)

# Supported file extensions and their types
_SUPPORTED_EXTENSIONS = {
    ".csv": "csv",
    ".tsv": "csv",
    ".json": "json",
    ".jsonl": "jsonl",
    ".ndjson": "jsonl",
    ".parquet": "parquet",
    ".xlsx": "excel",
    ".xls": "excel",
}


class FileStorageAdapter(ResourceAdapter):
    """Adapter for scanning files: CSV, JSON, Parquet, Excel.

    ConnectionConfig usage:
        uri: Directory path or single file path
        params:
            glob: Glob pattern for file discovery (default: all supported types)
            recursive: Whether to search recursively (default: True)
            csv_delimiter: CSV delimiter (default: auto-detect)
            csv_encoding: CSV encoding (default: "utf-8")
            json_array_key: Key for JSON array root (default: auto-detect)
            max_file_size_mb: Skip files larger than this (default: 100)

    Example:
        >>> resource = DataResource(
        ...     name="data-files",
        ...     resource_type=ResourceType.FILE_STORAGE,
        ...     connection=ConnectionConfig(
        ...         uri="/path/to/data",
        ...         params={"glob": "**/*.csv"},
        ...     ),
        ... )
        >>> with FileStorageAdapter(resource) as adapter:
        ...     files = adapter.list_containers()
    """

    def __init__(self, resource: DataResource) -> None:
        super().__init__(resource)
        self._base_path: Optional[Path] = None
        self._files: List[Path] = []

    def connect(self) -> None:
        """Validate path and discover files."""
        uri = self.resource.connection.uri
        if not uri:
            raise ConnectionError("File storage URI (path) is required.")

        self._base_path = Path(uri)
        if not self._base_path.exists():
            raise ConnectionError(f"Path does not exist: {self._base_path}")

        # Discover files
        params = self.resource.connection.params
        glob_pattern = params.get("glob")
        recursive = params.get("recursive", True)
        max_size_mb = params.get("max_file_size_mb", 100)
        max_size_bytes = max_size_mb * 1024 * 1024

        if self._base_path.is_file():
            self._files = [self._base_path]
        else:
            if glob_pattern:
                if recursive:
                    self._files = sorted(self._base_path.glob(glob_pattern))
                else:
                    self._files = sorted(self._base_path.glob(glob_pattern))
            else:
                # Discover all supported file types
                self._files = []
                for ext in _SUPPORTED_EXTENSIONS:
                    pattern = f"**/*{ext}" if recursive else f"*{ext}"
                    self._files.extend(self._base_path.glob(pattern))
                self._files = sorted(set(self._files))

        # Filter by size and supported extensions
        self._files = [
            f
            for f in self._files
            if f.is_file()
            and f.suffix.lower() in _SUPPORTED_EXTENSIONS
            and f.stat().st_size <= max_size_bytes
        ]

        self._connected = True
        logger.info(
            f"File storage connected: {self.resource.name}, " f"{len(self._files)} files found"
        )

    def close(self) -> None:
        """Release file references."""
        self._files = []
        self._base_path = None
        self._connected = False

    def list_containers(
        self,
        pattern: Optional[str] = None,
    ) -> List[ContainerInfo]:
        """List discovered files as containers.

        Args:
            pattern: Optional glob pattern to further filter files.

        Returns:
            List of ContainerInfo, one per file.
        """
        self._ensure_connected()

        containers: List[ContainerInfo] = []
        for file_path in self._files:
            name = (
                str(file_path.relative_to(self._base_path))
                if self._base_path and not self._base_path.is_file()
                else file_path.name
            )

            if pattern and not fnmatch.fnmatch(name, pattern):
                continue

            file_type = _SUPPORTED_EXTENSIONS.get(file_path.suffix.lower(), "unknown")
            stat = file_path.stat()

            containers.append(
                ContainerInfo(
                    name=name,
                    container_type=ContainerType.FILE,
                    metadata={
                        "file_type": file_type,
                        "file_path": str(file_path),
                        "size_bytes": stat.st_size,
                    },
                )
            )

        return containers

    def list_fields(self, container_name: str) -> List[FieldInfo]:
        """List columns/keys from a file.

        - CSV/TSV: First row headers
        - JSON: Top-level keys or keys from first array element
        - JSONL: Keys from first line
        - Parquet: Schema columns
        - Excel: First row headers

        Args:
            container_name: File name (as returned by list_containers).

        Returns:
            List of FieldInfo objects.
        """
        self._ensure_connected()

        file_path = self._resolve_file(container_name)
        if file_path is None:
            return []

        file_type = _SUPPORTED_EXTENSIONS.get(file_path.suffix.lower(), "unknown")

        try:
            if file_type == "csv":
                return self._list_csv_fields(file_path, container_name)
            elif file_type == "json":
                return self._list_json_fields(file_path, container_name)
            elif file_type == "jsonl":
                return self._list_jsonl_fields(file_path, container_name)
            elif file_type == "parquet":
                return self._list_parquet_fields(file_path, container_name)
            elif file_type == "excel":
                return self._list_excel_fields(file_path, container_name)
        except Exception as e:
            logger.warning(f"Could not list fields for {container_name}: {e}")

        return []

    def sample_values(
        self,
        container_name: str,
        field_name: str,
        limit: int = 100,
    ) -> List[str]:
        """Read sample values from a file column/key.

        Args:
            container_name: File name.
            field_name: Column or key name.
            limit: Maximum values.

        Returns:
            List of string values.
        """
        self._ensure_connected()

        file_path = self._resolve_file(container_name)
        if file_path is None:
            return []

        file_type = _SUPPORTED_EXTENSIONS.get(file_path.suffix.lower(), "unknown")

        try:
            if file_type == "csv":
                return self._sample_csv(file_path, field_name, limit)
            elif file_type == "json":
                return self._sample_json(file_path, field_name, limit)
            elif file_type == "jsonl":
                return self._sample_jsonl(file_path, field_name, limit)
            elif file_type == "parquet":
                return self._sample_parquet(file_path, field_name, limit)
            elif file_type == "excel":
                return self._sample_excel(file_path, field_name, limit)
        except Exception as e:
            logger.warning(f"Could not sample {field_name} from {container_name}: {e}")

        return []

    # ── File resolution ──

    def _resolve_file(self, container_name: str) -> Optional[Path]:
        """Resolve container name to file path."""
        for file_path in self._files:
            if self._base_path and not self._base_path.is_file():
                relative = str(file_path.relative_to(self._base_path))
            else:
                relative = file_path.name
            if relative == container_name:
                return file_path
        return None

    # ── CSV ──

    def _list_csv_fields(self, path: Path, container_name: str) -> List[FieldInfo]:
        encoding = self.resource.connection.params.get("csv_encoding", "utf-8")
        delimiter = self.resource.connection.params.get("csv_delimiter")

        with open(path, encoding=encoding, newline="") as f:
            if delimiter is None:
                sample = f.read(8192)
                try:
                    dialect = csv.Sniffer().sniff(sample)
                    delimiter = dialect.delimiter
                except csv.Error:
                    delimiter = ","
                f.seek(0)

            reader = csv.reader(f, delimiter=delimiter)
            headers = next(reader, None)

        if not headers:
            return []

        return [
            FieldInfo(
                name=h.strip(),
                container_name=container_name,
                data_type="string",
                metadata={"file_type": "csv"},
            )
            for h in headers
            if h.strip()
        ]

    def _sample_csv(self, path: Path, field_name: str, limit: int) -> List[str]:
        encoding = self.resource.connection.params.get("csv_encoding", "utf-8")
        delimiter = self.resource.connection.params.get("csv_delimiter", ",")

        values: List[str] = []
        with open(path, encoding=encoding, newline="") as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            for row in reader:
                if len(values) >= limit:
                    break
                val = row.get(field_name)
                if val is not None and val.strip():
                    values.append(val.strip())

        return values

    # ── JSON ──

    def _list_json_fields(self, path: Path, container_name: str) -> List[FieldInfo]:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)

        # If array, use first element
        sample = data
        array_key = self.resource.connection.params.get("json_array_key")
        if array_key and isinstance(data, dict):
            sample = data.get(array_key, data)
        if isinstance(sample, list) and sample:
            sample = sample[0]
        if not isinstance(sample, dict):
            return []

        return [
            FieldInfo(
                name=key,
                container_name=container_name,
                data_type=type(val).__name__,
                metadata={"file_type": "json"},
            )
            for key, val in sample.items()
        ]

    def _sample_json(self, path: Path, field_name: str, limit: int) -> List[str]:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)

        records = data
        array_key = self.resource.connection.params.get("json_array_key")
        if array_key and isinstance(data, dict):
            records = data.get(array_key, [])
        if isinstance(records, dict):
            records = [records]
        if not isinstance(records, list):
            return []

        values: List[str] = []
        for record in records[:limit]:
            if isinstance(record, dict):
                val = record.get(field_name)
                if val is not None:
                    values.append(str(val))

        return values

    # ── JSONL ──

    def _list_jsonl_fields(self, path: Path, container_name: str) -> List[FieldInfo]:
        with open(path, encoding="utf-8") as f:
            first_line = f.readline().strip()

        if not first_line:
            return []

        try:
            record = json.loads(first_line)
        except json.JSONDecodeError:
            return []

        if not isinstance(record, dict):
            return []

        return [
            FieldInfo(
                name=key,
                container_name=container_name,
                data_type=type(val).__name__,
                metadata={"file_type": "jsonl"},
            )
            for key, val in record.items()
        ]

    def _sample_jsonl(self, path: Path, field_name: str, limit: int) -> List[str]:
        values: List[str] = []
        with open(path, encoding="utf-8") as f:
            for line in f:
                if len(values) >= limit:
                    break
                line = line.strip()
                if not line:
                    continue
                try:
                    record = json.loads(line)
                    if isinstance(record, dict):
                        val = record.get(field_name)
                        if val is not None:
                            values.append(str(val))
                except json.JSONDecodeError:
                    continue

        return values

    # ── Parquet ──

    def _list_parquet_fields(self, path: Path, container_name: str) -> List[FieldInfo]:
        try:
            import pyarrow.parquet as pq
        except ImportError:
            raise ImportError(
                "pyarrow is required for Parquet file scanning. "
                "Install with: pip install data-detector[file-storage]"
            )

        schema = pq.read_schema(str(path))
        return [
            FieldInfo(
                name=field.name,
                container_name=container_name,
                data_type=str(field.type),
                nullable=field.nullable,
                metadata={"file_type": "parquet"},
            )
            for field in schema
        ]

    def _sample_parquet(self, path: Path, field_name: str, limit: int) -> List[str]:
        try:
            import pyarrow.parquet as pq
        except ImportError:
            raise ImportError(
                "pyarrow is required for Parquet file scanning. "
                "Install with: pip install data-detector[file-storage]"
            )

        table = pq.read_table(str(path), columns=[field_name])
        column = table.column(field_name)
        values: List[str] = []
        for val in column.to_pylist()[:limit]:
            if val is not None:
                values.append(str(val))
        return values

    # ── Excel ──

    def _list_excel_fields(self, path: Path, container_name: str) -> List[FieldInfo]:
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel file scanning. "
                "Install with: pip install data-detector[file-storage]"
            )

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        fields: List[FieldInfo] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(h) if h is not None else "" for h in row]
                break

            for h in headers:
                if h.strip():
                    fields.append(
                        FieldInfo(
                            name=h.strip(),
                            container_name=container_name,
                            data_type="string",
                            metadata={"file_type": "excel", "sheet": sheet_name},
                        )
                    )

        wb.close()
        return fields

    def _sample_excel(self, path: Path, field_name: str, limit: int) -> List[str]:
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel file scanning. "
                "Install with: pip install data-detector[file-storage]"
            )

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        values: List[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            col_idx = None

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    # Find column index
                    for i, h in enumerate(row):
                        if h is not None and str(h).strip() == field_name:
                            col_idx = i
                            break
                    if col_idx is None:
                        break
                    continue

                if len(values) >= limit:
                    break

                if col_idx < len(row) and row[col_idx] is not None:
                    values.append(str(row[col_idx]))

            if values:
                break

        wb.close()
        return values
